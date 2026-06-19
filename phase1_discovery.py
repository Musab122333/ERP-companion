"""
Phase 1 – ERP Discovery & Data Dictionary
==========================================
Run this on your Windows machine where DBF files are stored.

Install dependencies first:
    pip install dbfread openpyxl pandas

Usage:
    python phase1_discovery.py

Output:
    data_dictionary.xlsx   — Full schema + samples + relationship map
    discovery_report.json  — Machine-readable schema for Phase 2
"""

import os
import json
import struct
import hashlib
import traceback
from pathlib import Path
from datetime import datetime
from collections import defaultdict

# ---------------------------------------------------------------------------
# CONFIG — edit these paths to match your system
# ---------------------------------------------------------------------------
ERP_ROOT_FOLDERS = [
    r"D:\FA\ZF2526",
    r"D:\FA\ZF2627",
    r"D:\FA\ZFACA2627",
    r"D:\FA\ZFN2526",
    # Add more folders here as needed
]

OUTPUT_DIR = r"C:\ERP_Discovery"
SAMPLE_ROWS = 5          # How many sample rows to capture per table
MAX_FIELDS_SHOWN = 50    # Safety cap for very wide tables

# ---------------------------------------------------------------------------
# DBF READER — pure Python, no external dependency fallback
# ---------------------------------------------------------------------------

DBF_FIELD_TYPES = {
    'C': 'Character',
    'N': 'Numeric',
    'F': 'Float',
    'D': 'Date',
    'L': 'Logical',
    'M': 'Memo',
    'B': 'Binary',
    'I': 'Integer',
    'Y': 'Currency',
    'T': 'DateTime',
    'G': 'General',
}

import re as _re

def sanitize(value):
    """Remove characters that are illegal in Excel cells (openpyxl)."""
    if value is None:
        return ''
    s = str(value)
    # Strip null bytes and non-printable control chars (except tab/newline)
    s = s.replace('\x00', '')
    s = _re.sub(r'[\x01-\x08\x0b\x0c\x0e-\x1f\x7f\ufffd]', '', s)
    return s.strip()

def read_dbf_header(filepath):
    """Read DBF header and return (fields, record_count) without external libs."""
    fields = []
    record_count = 0
    try:
        with open(filepath, 'rb') as f:
            header = f.read(32)
            if len(header) < 32:
                return fields, record_count

            record_count = struct.unpack_from('<I', header, 4)[0]
            header_size = struct.unpack_from('<H', header, 8)[0]
            record_size = struct.unpack_from('<H', header, 10)[0]

            # Read field descriptors (32 bytes each)
            while True:
                field_data = f.read(32)
                if not field_data or field_data[0] == 0x0D:
                    break
                if len(field_data) < 32:
                    break

                raw_name = field_data[:11].replace(b'\x00', b'')
                # Decode strictly to ASCII; drop any non-ASCII bytes entirely
                field_name = raw_name.decode('ascii', errors='ignore').strip()
                # Sanitize away any remaining illegal Excel characters
                field_name = sanitize(field_name)

                type_byte = field_data[11]
                field_type = chr(type_byte) if 0x20 <= type_byte <= 0x7E else '?'
                field_length = field_data[16]
                field_decimal = field_data[17]

                if field_name:
                    fields.append({
                        'name': field_name,
                        'type': field_type,
                        'type_desc': DBF_FIELD_TYPES.get(field_type, f'Unknown({field_type})'),
                        'length': field_length,
                        'decimal': field_decimal,
                    })

    except Exception as e:
        print(f"    [WARN] Could not read header: {e}")
    return fields, record_count


def read_dbf_sample(filepath, fields, max_rows=5):
    """Read sample rows from DBF. Returns list of dicts."""
    rows = []
    try:
        # Try dbfread first (best option)
        try:
            from dbfread import DBF
            table = DBF(filepath, encoding='cp1252', ignore_missing_memofile=True)
            for i, record in enumerate(table):
                if i >= max_rows:
                    break
                row = {}
                for k, v in record.items():
                    row[str(k)] = str(v) if v is not None else ''
                rows.append(row)
            return rows
        except ImportError:
            pass

        # Fallback: manual binary read
        with open(filepath, 'rb') as f:
            header = f.read(32)
            if len(header) < 32:
                return rows
            header_size = struct.unpack_from('<H', header, 8)[0]
            record_size = struct.unpack_from('<H', header, 10)[0]
            record_count = struct.unpack_from('<I', header, 4)[0]

            f.seek(header_size)
            total_field_len = sum(f['length'] for f in fields)

            for i in range(min(max_rows, record_count)):
                raw = f.read(record_size)
                if not raw or len(raw) < record_size:
                    break
                if raw[0] == 0x2A:  # deleted record
                    continue

                row = {}
                offset = 1  # skip deletion flag
                for field in fields:
                    chunk = raw[offset:offset + field['length']]
                    try:
                        value = chunk.decode('cp1252', errors='ignore').strip()
                    except Exception:
                        value = ''
                    row[field['name']] = sanitize(value)
                    offset += field['length']
                rows.append(row)

    except Exception as e:
        print(f"    [WARN] Could not read sample rows: {e}")
    return rows


# ---------------------------------------------------------------------------
# DISCOVERY ENGINE
# ---------------------------------------------------------------------------

def discover_folder(folder_path):
    """Scan one ERP folder and return discovery data."""
    folder = Path(folder_path)
    if not folder.exists():
        print(f"  [SKIP] Folder not found: {folder_path}")
        return []

    dbf_files = sorted(folder.glob("*.DBF")) + sorted(folder.glob("*.dbf"))
    results = []

    for dbf_path in dbf_files:
        table_name = dbf_path.stem.upper()
        print(f"  → Reading {table_name} ...", end="", flush=True)

        fields, record_count = read_dbf_header(str(dbf_path))
        sample_rows = read_dbf_sample(str(dbf_path), fields, max_rows=SAMPLE_ROWS)

        file_stat = dbf_path.stat()

        result = {
            'source_folder': folder.name,
            'full_path': str(dbf_path),
            'table_name': table_name,
            'record_count': record_count,
            'field_count': len(fields),
            'file_size_kb': round(file_stat.st_size / 1024, 1),
            'last_modified': datetime.fromtimestamp(file_stat.st_mtime).strftime('%Y-%m-%d %H:%M'),
            'fields': fields,
            'sample_rows': sample_rows,
        }
        results.append(result)
        print(f" {record_count:,} records, {len(fields)} fields ✓")

    return results


def find_relationships(all_tables):
    """Detect fields that appear in multiple tables — likely join keys."""
    field_map = defaultdict(list)  # field_name -> list of (source, table)

    for tbl in all_tables:
        for field in tbl['fields']:
            key = field['name'].upper()
            field_map[key].append(f"{tbl['source_folder']}/{tbl['table_name']}")

    relationships = {
        field: tables
        for field, tables in field_map.items()
        if len(tables) > 1
    }
    return relationships


# ---------------------------------------------------------------------------
# EXCEL REPORT BUILDER
# ---------------------------------------------------------------------------

def build_excel_report(all_tables, relationships, output_path):
    """Create a well-formatted Excel data dictionary."""
    from openpyxl import Workbook
    from openpyxl.styles import (
        Font, PatternFill, Alignment, Border, Side
    )
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    # Color palette
    COLOR_HEADER    = "1F4E79"  # dark blue
    COLOR_SUBHEADER = "2E75B6"  # medium blue
    COLOR_ALT_ROW   = "DEEAF1"  # light blue
    COLOR_YELLOW    = "FFF2CC"
    COLOR_GREEN     = "E2EFDA"
    COLOR_WHITE     = "FFFFFF"

    def style_header(cell, bg=COLOR_HEADER, font_color="FFFFFF", bold=True, size=11):
        cell.font = Font(bold=bold, color=font_color, size=size, name='Arial')
        cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    def style_cell(cell, bold=False, bg=None, align='left'):
        cell.font = Font(bold=bold, size=10, name='Arial')
        if bg:
            cell.fill = PatternFill("solid", start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center')

    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    def apply_border(ws, min_row, max_row, min_col, max_col):
        for row in ws.iter_rows(min_row=min_row, max_row=max_row,
                                min_col=min_col, max_col=max_col):
            for cell in row:
                cell.border = border

    # ── Sheet 1: Summary ──────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.row_dimensions[1].height = 30

    headers = ["Source Folder", "Table Name", "Record Count", "Field Count",
               "File Size (KB)", "Last Modified", "Known Purpose"]
    for col, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col, value=h)
        style_header(cell)

    KNOWN_PURPOSE = {
        'ITEM':   'Product / Item Master',
        'MASTER': 'Ledger Master (Customers, Suppliers, Accounts)',
        'SALES1': 'Sales Invoice Header',
        'SALES2': 'Sales Invoice Line Items',
        'TRANS':  'Accounting Transactions / Vouchers',
    }

    for i, tbl in enumerate(all_tables):
        row = i + 2
        bg = COLOR_ALT_ROW if i % 2 == 0 else COLOR_WHITE
        values = [
            tbl['source_folder'],
            tbl['table_name'],
            tbl['record_count'],
            tbl['field_count'],
            tbl['file_size_kb'],
            tbl['last_modified'],
            KNOWN_PURPOSE.get(tbl['table_name'], '— To be analysed'),
        ]
        for col, val in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col, value=val)
            style_cell(cell, bg=bg)
            if col == 3:  # record count
                cell.number_format = '#,##0'
                cell.alignment = Alignment(horizontal='right')

    # Totals row
    total_row = len(all_tables) + 2
    ws_summary.cell(row=total_row, column=1, value="TOTAL").font = Font(bold=True, name='Arial')
    total_recs = ws_summary.cell(row=total_row, column=3,
                                  value=f"=SUM(C2:C{total_row-1})")
    total_recs.font = Font(bold=True, name='Arial')
    total_recs.number_format = '#,##0'
    total_tbls = ws_summary.cell(row=total_row, column=2,
                                  value=f"=COUNTA(B2:B{total_row-1})")
    total_tbls.font = Font(bold=True, name='Arial')

    apply_border(ws_summary, 1, total_row, 1, len(headers))

    col_widths = [18, 16, 15, 13, 15, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws_summary.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 2: Field Dictionary ─────────────────────────────────────────
    ws_fields = wb.create_sheet("Field Dictionary")
    ws_fields.row_dimensions[1].height = 28

    fheaders = ["Source Folder", "Table", "Field Name", "Type",
                "Type Description", "Length", "Decimal",
                "Likely Meaning", "Is Join Key?"]
    for col, h in enumerate(fheaders, 1):
        cell = ws_fields.cell(row=1, column=col, value=h)
        style_header(cell)

    # Known field meanings
    FIELD_MEANINGS = {
        'INV_NO':  'Invoice Number', 'INVNO': 'Invoice Number',
        'INV_DATE':'Invoice Date',   'INVDATE': 'Invoice Date',
        'ITEM_CODE':'Item Code',     'ITEMCODE': 'Item Code',
        'CUST_CODE':'Customer Code', 'CUSTCODE': 'Customer Code',
        'QTY':     'Quantity',       'RATE': 'Rate / Unit Price',
        'AMOUNT':  'Line Amount',    'DISC': 'Discount',
        'BILL_AMT':'Bill Amount',    'TAX': 'Tax Amount',
        'LEDCODE': 'Ledger Code',    'NARR': 'Narration',
        'VCH_NO':  'Voucher Number', 'VCH_DATE': 'Voucher Date',
        'DR_CR':   'Debit/Credit',   'DRCR': 'Debit/Credit Flag',
        'HSN':     'HSN Code',       'UOM': 'Unit of Measure',
        'OPENING': 'Opening Balance','CLOSING': 'Closing Balance',
        'MOBILE':  'Mobile Number',  'STATE': 'State',
        'GROUP':   'Account Group',
    }

    join_key_fields = set(relationships.keys())
    field_row = 2

    for tbl in all_tables:
        prev_table = None
        for j, field in enumerate(tbl['fields']):
            bg = COLOR_ALT_ROW if j % 2 == 0 else COLOR_WHITE
            is_join = "✓ YES" if field['name'].upper() in join_key_fields else ""
            meaning = FIELD_MEANINGS.get(field['name'].upper(), "")

            row_vals = [
                sanitize(tbl['source_folder']),
                sanitize(tbl['table_name']),
                sanitize(field['name']),
                sanitize(field['type']),
                sanitize(field['type_desc']),
                field['length'],
                field['decimal'] if field['decimal'] else '',
                sanitize(meaning),
                sanitize(is_join),
            ]
            for col, val in enumerate(row_vals, 1):
                cell = ws_fields.cell(row=field_row, column=col, value=val)
                style_cell(cell, bg=bg)
                if is_join and col == 9:
                    cell.font = Font(bold=True, color="375623", name='Arial')
                    cell.fill = PatternFill("solid", start_color=COLOR_GREEN)
            field_row += 1

    apply_border(ws_fields, 1, field_row - 1, 1, len(fheaders))
    fw = [18, 14, 18, 8, 18, 8, 10, 28, 12]
    for i, w in enumerate(fw, 1):
        ws_fields.column_dimensions[get_column_letter(i)].width = w

    # ── Sheet 3: Relationship Map ─────────────────────────────────────────
    ws_rel = wb.create_sheet("Relationship Map")
    ws_rel.row_dimensions[1].height = 28
    ws_rel.cell(row=1, column=1, value="Shared Field (Potential Join Key)")
    ws_rel.cell(row=1, column=2, value="Tables That Contain This Field")
    ws_rel.cell(row=1, column=3, value="Join Type Guess")
    for col in range(1, 4):
        style_header(ws_rel.cell(row=1, column=col))

    JOIN_GUESSES = {
        'INV_NO':  'SALES1.INV_NO = SALES2.INV_NO  (Header ↔ Lines)',
        'INVNO':   'SALES1.INVNO  = SALES2.INVNO   (Header ↔ Lines)',
        'ITEM_CODE':'SALES2.ITEM_CODE = ITEM.ITEM_CODE (Lines ↔ Product)',
        'ITEMCODE': 'SALES2.ITEMCODE  = ITEM.ITEMCODE  (Lines ↔ Product)',
        'CUST_CODE':'SALES1.CUST_CODE = MASTER.CODE  (Invoice ↔ Ledger)',
        'LEDCODE':  'TRANS.LEDCODE    = MASTER.CODE  (Trans ↔ Ledger)',
    }

    for i, (field, tables) in enumerate(sorted(relationships.items())):
        row = i + 2
        bg = COLOR_ALT_ROW if i % 2 == 0 else COLOR_WHITE
        ws_rel.cell(row=row, column=1, value=field)
        ws_rel.cell(row=row, column=2, value=", ".join(tables))
        ws_rel.cell(row=row, column=3, value=JOIN_GUESSES.get(field.upper(), "— Analyse manually"))
        for col in range(1, 4):
            style_cell(ws_rel.cell(row=row, column=col), bg=bg)

    apply_border(ws_rel, 1, len(relationships) + 1, 1, 3)
    for i, w in enumerate([30, 60, 50], 1):
        ws_rel.column_dimensions[get_column_letter(i)].width = w

    # ── Sheets 4+: Sample Data per table ──────────────────────────────────
    for tbl in all_tables:
        if not tbl['sample_rows']:
            continue
        safe_name = f"{tbl['source_folder'][:6]}_{tbl['table_name']}"[:31]
        ws_s = wb.create_sheet(safe_name)

        # Title row
        ws_s.merge_cells(start_row=1, start_column=1,
                          end_row=1, end_column=max(len(tbl['fields']), 1))
        title_cell = ws_s.cell(row=1, column=1,
                                value=f"{tbl['source_folder']} / {tbl['table_name']}  "
                                      f"({tbl['record_count']:,} records)")
        style_header(title_cell, size=12)
        ws_s.row_dimensions[1].height = 24

        # Column headers
        for col, field in enumerate(tbl['fields'][:MAX_FIELDS_SHOWN], 1):
            cell = ws_s.cell(row=2, column=col, value=sanitize(field['name']))
            style_header(cell, bg=COLOR_SUBHEADER)
            ws_s.column_dimensions[get_column_letter(col)].width = max(
                len(field['name']) + 2, 12)

        # Sample data
        for r, row_data in enumerate(tbl['sample_rows'], 3):
            bg = COLOR_ALT_ROW if r % 2 == 0 else COLOR_WHITE
            for col, field in enumerate(tbl['fields'][:MAX_FIELDS_SHOWN], 1):
                val = sanitize(row_data.get(field['name'], ''))
                cell = ws_s.cell(row=r, column=col, value=val)
                style_cell(cell, bg=bg)

        apply_border(ws_s, 2, 2 + len(tbl['sample_rows']), 1,
                     min(len(tbl['fields']), MAX_FIELDS_SHOWN))

    wb.save(output_path)
    print(f"\n✅  Excel report saved: {output_path}")


# ---------------------------------------------------------------------------
# JSON REPORT (for Phase 2 sync engine)
# ---------------------------------------------------------------------------

def build_json_report(all_tables, relationships, output_path):
    report = {
        'generated_at': datetime.now().isoformat(),
        'total_tables': len(all_tables),
        'total_records': sum(t['record_count'] for t in all_tables),
        'relationships': relationships,
        'tables': []
    }

    for tbl in all_tables:
        report['tables'].append({
            'source_folder': tbl['source_folder'],
            'table_name': tbl['table_name'],
            'full_path': tbl['full_path'],
            'record_count': tbl['record_count'],
            'fields': tbl['fields'],
        })

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(report, f, indent=2, default=str)
    print(f"✅  JSON report saved:  {output_path}")


# ---------------------------------------------------------------------------
# MAIN
# ---------------------------------------------------------------------------

def main():
    print("=" * 60)
    print("  ERP Discovery Tool — Phase 1")
    print(f"  {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print("=" * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    all_tables = []
    for folder in ERP_ROOT_FOLDERS:
        folder_name = Path(folder).name
        print(f"\n📂 Scanning: {folder}")
        tables = discover_folder(folder)
        all_tables.extend(tables)

    if not all_tables:
        print("\n⚠️  No DBF files found. Check ERP_ROOT_FOLDERS paths in config.")
        return

    print(f"\n🔍 Analysing relationships across {len(all_tables)} tables ...")
    relationships = find_relationships(all_tables)
    print(f"   Found {len(relationships)} shared fields (potential join keys)")

    xlsx_path = os.path.join(OUTPUT_DIR, "data_dictionary.xlsx")
    json_path = os.path.join(OUTPUT_DIR, "discovery_report.json")

    print(f"\n📊 Building Excel data dictionary ...")
    build_excel_report(all_tables, relationships, xlsx_path)

    print(f"📄 Building JSON schema report ...")
    build_json_report(all_tables, relationships, json_path)

    # Print quick summary
    print("\n" + "=" * 60)
    print("  DISCOVERY SUMMARY")
    print("=" * 60)
    print(f"  Folders scanned : {len(ERP_ROOT_FOLDERS)}")
    print(f"  Tables found    : {len(all_tables)}")
    total_records = sum(t['record_count'] for t in all_tables)
    print(f"  Total records   : {total_records:,}")
    print(f"  Shared fields   : {len(relationships)}")
    print(f"\n  Outputs in      : {OUTPUT_DIR}")
    print("=" * 60)
    print("\n  Next step → Phase 2: Sync Engine")
    print("  Share discovery_report.json to begin schema design.\n")


if __name__ == "__main__":
    main()